import os

from openpyxl import load_workbook, Workbook
from dateutil.parser import parse
from datetime import timedelta, datetime
from pathlib import Path
import time


class ExcelLogic:

    def __init__(self, file_path):
        self.filePath = file_path
        self.ws = None
        self.preview = None

    def load_preview(self, file_name, n=250):
        wb = load_workbook(file_name, read_only=True)
        self.ws = wb.active
        return_prev = []
        for row_idx, row in enumerate(self.ws.iter_rows()):
            row = [x.value for x in row]
            if row_idx == 0:
                return_prev.append(self.format_column_names(row))
                continue
            return_prev.append(row)
            if row_idx >= n:
                break
        self.preview = return_prev

    @classmethod
    def format_column_names(cls, cols):
        cols_reps = {}
        final_cols = []
        for col in cols:
            if col is None:
                col = "NULL"
            if col not in final_cols:
                final_cols.append(col)
                cols_reps[col] = 0
            else:
                cols_reps[col] += 1
                col += "." + str(cols_reps[col])
                final_cols.append(col)
        return final_cols

    def detect_ensay_start(self, master, threshold=10):
        columns = None
        previous_val = None
        col_index = None
        first_data = None
        for row_idx, row in enumerate(self.ws.iter_rows()):
            row = [x.value for x in row]
            if row_idx == 0:
                columns = self.format_column_names(row)
                col_index = columns.index(master)
                continue
            if row_idx == 1:
                first_data = row
            if previous_val is None:
                previous_val = row[col_index]
                continue
            if abs(row[col_index] - previous_val) > threshold:
                return row, columns, first_data
            previous_val = row[col_index]
        return None, None, None

    def generate_ensays_data(self, ensay_rows, total_rows_to_iter, signal):
        ensay_results = {x: [] for x in range(len(ensay_rows))}
        for row_idx, row in enumerate(self.ws.iter_rows()):
            if row_idx == 0:
                continue
            row = [x.value for x in row]
            rd = row[0]
            #  year, month=None, day=None, hour=0, minute=0, second=0
            # cur_date = datetime(int(rd[0:4]), int(rd[5:7]), int(rd[8:10]), int(rd[11:13]), int(rd[14:16]), int(rd[17:19]))
            cur_date = rd
            if not isinstance(rd, datetime):
                cur_date = parse(rd)

            total_ends = 0
            for ensay_idx, erow in enumerate(ensay_rows):
                if erow[0] <= cur_date <= erow[1]:
                    ensay_results[ensay_idx].append(row)
                    continue
                if cur_date > erow[1]:
                    total_ends += 1
            if total_ends >= len(ensay_rows):
                break
            signal.emit([int((row_idx + 1) / total_rows_to_iter * 100), ensay_rows])
        return ensay_results

    def calculate_ensys(self, master_column, duration, offset, num_ensays, out_name, signal):
        ensay_start, column_names, first_data = self.detect_ensay_start(master_column)
        ensay_start_date = ensay_start[0]
        if not isinstance(ensay_start[0], datetime):
            ensay_start_date = parse(ensay_start[0])
        ensays_dates = []
        for idx, num in enumerate(range(1, num_ensays + 1)):
            st_wo_offset = ensay_start_date + timedelta(minutes=(duration * (num - 1)))
            start_date = st_wo_offset - timedelta(minutes=offset)
            start_date = parse(start_date.strftime("%Y-%m-%d %H:%M:00"))
            end_date = st_wo_offset + timedelta(minutes=(duration + offset))
            end_date = parse(end_date.strftime("%Y-%m-%d %H:%M:00"))
            ensays_dates.append((start_date, end_date, st_wo_offset))

        first_data_time = first_data[0]
        if not isinstance(first_data_time, datetime):
            first_data_time = parse(first_data_time)
        z_time = time.mktime(first_data_time.timetuple())

        st_time = time.mktime(ensays_dates[0][0].timetuple())
        nd_time = time.mktime(ensays_dates[0][1].timetuple())
        dur = int(nd_time - st_time)
        z_dur = int(st_time - z_time)

        total_rows_to_iter = (z_dur * 4) + (dur * 4 * num_ensays)

        ensay_results = self.generate_ensays_data(ensays_dates, total_rows_to_iter, signal)

        Path("ensay_outputs").mkdir(exist_ok=True, parents=True)
        results_statics = []
        for eidx, edr in ensay_results.items():
            wb = Workbook(write_only=True)
            ws = wb.create_sheet()
            ws.append(column_names)
            for dr in edr:
                ws.append(dr)
            wb.save("ensay_outputs" + os.sep + f"{out_name}_{eidx + 1}.xlsx")
            results_statics.append([ensays_dates[eidx][0], ensays_dates[eidx][1], ensays_dates[eidx][2], len(edr)])
            signal.emit([int((eidx + 1) / len(ensay_results) * 100), ensays_dates])
        return results_statics
